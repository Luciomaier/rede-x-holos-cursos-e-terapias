#!/usr/bin/env python3
"""Lê a planilha META DE VENDAS de forma confiável (todas as abas, todas as linhas).

Por que existe: a planilha tem FILTRO ATIVO. A coluna QUANT. respeita o filtro e subconta;
exports por CSV/gviz/conector do Drive omitem as linhas filtradas. Só o .xlsx completo traz tudo.
Foi assim que junho/2026 foi reportado com 27 matrículas quando eram 32.

Uso:
    # 1. baixar o xlsx (navegador logado):
    #    https://docs.google.com/spreadsheets/d/1MYA0OhQSBjKDr3hsL-aDK0Vv-_dxCp-0T353E2UIpuU/export?format=xlsx
    python3 ler-meta-de-vendas.py META-DE-VENDAS.xlsx jun26

Saída: vendas de masso por dezena e por categoria + a CONFERÊNCIA obrigatória
(soma do detalhe == VENDIDO do cabeçalho). Se não bater, tem linha oculta.
"""
import sys
import openpyxl

MASSO = {
    "1ª mensalidade": "mensalista",
    "massoterapia integral": "integral",
    "massoterapia intensivo": "intensivo",
}
# O valor da venda pode estar em qualquer um dos 4 baldes de comissão — somar os quatro.
COLS_VALOR = (6, 7, 8, 9)  # com. 1 | com. 2 | escola | Cris


def num(v):
    """Célula da planilha → float. Vazio, '-' e texto viram 0."""
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def brl(n):
    return f"R$ {n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def dezena(dia):
    return "D1" if dia <= 10 else ("D2" if dia <= 20 else "D3")


def main(path, aba):
    ws = openpyxl.load_workbook(path, data_only=True)[aba]

    cabecalho = {}
    vendas = []
    for r in range(1, ws.max_row + 1):
        rotulo = ws.cell(r, 3).value  # coluna C: nome do produto no bloco de cima
        if rotulo and "Massoterapia" in str(rotulo):
            cabecalho[str(rotulo).strip()] = num(ws.cell(r, 7).value)  # coluna G = VENDIDO

        curso = ws.cell(r, 5).value  # coluna E: curso, no bloco de detalhe
        cat = MASSO.get(str(curso).strip().lower()) if curso else None
        if not cat:
            continue
        data = ws.cell(r, 1).value
        if not hasattr(data, "day"):
            print(f"  ⚠️  L{r} sem data — fora da contagem: {ws.cell(r, 4).value}")
            continue
        vendas.append({
            "linha": r,
            "dia": data.day,
            "cat": cat,
            "valor": sum(num(ws.cell(r, c).value) for c in COLS_VALOR),
            "midia": (ws.cell(r, 2).value or "(vazio)").strip(),
            "nome": ws.cell(r, 4).value,
            "tel": ws.cell(r, 3).value,
        })

    print(f"\n=== {aba} — masso vendas NOVAS (detalhe lead-a-lead) ===\n")
    for k in ("D1", "D2", "D3"):
        sub = [v for v in vendas if dezena(v["dia"]) == k]
        if not sub:
            continue
        total = sum(v["valor"] for v in sub)
        cats = {}
        for v in sub:
            cats[v["cat"]] = cats.get(v["cat"], 0) + 1
        print(f"{k}: {len(sub):>2} vendas · {brl(total):>14} · ticket {brl(total/len(sub)):>12}   {cats}")

    n = len(vendas)
    v = sum(x["valor"] for x in vendas)
    print(f"\nMÊS: {n} vendas · {brl(v)} · ticket {brl(v/n)}")

    print("\n--- por categoria ---")
    for c in ("integral", "intensivo", "mensalista"):
        sub = [x for x in vendas if x["cat"] == c]
        print(f"  {c:<11} {len(sub):>2} · {brl(sum(x['valor'] for x in sub))}")

    # A conferência que impede o erro de junho/26 de acontecer de novo
    print("\n--- ⚠️  CONFERÊNCIA (detalhe × cabeçalho) ---")
    cab_total = sum(cabecalho.values())
    print(f"  cabeçalho (VENDIDO): {brl(cab_total)}")
    print(f"  detalhe (linhas):    {brl(v)}")
    dif = round(v - cab_total, 2)
    if abs(dif) < 0.01:
        print("  ✅ BATE — pode fechar o número.")
    else:
        print(f"  ❌ NÃO BATE — diferença de {brl(abs(dif))}.")
        print("     Tem linha oculta pelo filtro, ou valor numa coluna de comissão não somada.")
        print("     NÃO feche a dezena até resolver.")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    main(sys.argv[1], sys.argv[2])
